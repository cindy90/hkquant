"""
生成 NACS 数据收集模板 Excel

设计原则:
    - 每个 sheet 都有列头说明 + Wind指标 + 单位 + 必填/选填
    - 第2行是华勤(03296.HK)的示例数据, 含 Wind 公式实例供拖拽
    - 颜色编码:
        蓝色字 = 用户手填 (硬编码)
        黄色底 = Wind公式自动取数
        绿色字 = 派生公式
        灰色底 = 列头/说明
    - 关键字段加 data validation (下拉框)
    - README + 数据字典 + Wind公式速查表
"""
from __future__ import annotations
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                             NamedStyle)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule


# =============================================================================
# 样式常量
# =============================================================================

FONT_HEADER = Font(name="Arial", size=10, bold=True, color="FFFFFF")
FONT_USER_INPUT = Font(name="Arial", size=10, color="0000FF")        # 蓝
FONT_FORMULA = Font(name="Arial", size=10, color="000000")           # 黑
FONT_DERIVED = Font(name="Arial", size=10, color="008000")           # 绿
FONT_NOTE = Font(name="Arial", size=9, italic=True, color="666666")
FONT_TITLE = Font(name="Arial", size=14, bold=True, color="1F4E78")

FILL_HEADER = PatternFill("solid", start_color="1F4E78")             # 深蓝
FILL_SUBHEADER = PatternFill("solid", start_color="D9E1F2")          # 浅蓝
FILL_WIND = PatternFill("solid", start_color="FFF2CC")               # 浅黄(Wind公式)
FILL_REQUIRED = PatternFill("solid", start_color="FCE4D6")           # 浅橙(必填)
FILL_OPTIONAL = PatternFill("solid", start_color="E2EFDA")           # 浅绿(选填)
FILL_EXAMPLE = PatternFill("solid", start_color="F2F2F2")            # 浅灰(示例行)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

THIN = Side(border_style="thin", color="BFBFBF")
BORDER_ALL = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)


def apply_header_row(ws, row: int, headers: list, widths: list = None):
    """格式化列头行"""
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = FONT_HEADER
        c.fill = FILL_HEADER
        c.alignment = ALIGN_CENTER
        c.border = BORDER_ALL
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 38


def apply_subheader_row(ws, row: int, n_cols: int, content: list):
    """字段子说明行 (Wind指标/单位/必填)"""
    for i in range(1, n_cols + 1):
        c = ws.cell(row=row, column=i,
                    value=content[i-1] if i-1 < len(content) else "")
        c.font = FONT_NOTE
        c.fill = FILL_SUBHEADER
        c.alignment = ALIGN_CENTER
        c.border = BORDER_ALL
    ws.row_dimensions[row].height = 30


def apply_example_row(ws, row: int, n_cols: int):
    """示例行底色"""
    for i in range(1, n_cols + 1):
        c = ws.cell(row=row, column=i)
        if c.fill.start_color.rgb in (None, "00000000"):
            c.fill = FILL_EXAMPLE
        c.border = BORDER_ALL


# =============================================================================
# Sheet 1: README
# =============================================================================

def make_readme(wb: Workbook):
    ws = wb.create_sheet("00_README", 0)
    ws.column_dimensions["A"].width = 110

    rows = [
        ("NACS 数据收集模板 v1.0", FONT_TITLE, None),
        ("", None, None),
        ("用途", Font(name="Arial", size=12, bold=True, color="1F4E78"), None),
        ("本模板用于收集香港 IPO 基石投资量化模型 (NACS) 的回测原始数据。",
         None, None),
        ("覆盖范围: 2022-01-01 至 2026-04-30 期间在 HKEX 主板上市的全部新股 (~250只)。",
         None, None),
        ("",  None, None),
        ("工作流", Font(name="Arial", size=12, bold=True, color="1F4E78"), None),
        ("步骤 1: 在 sheet [01_ipo_master] 用 Wind WSET 函数批量拉出该期间所有 IPO 的代码列表。",
         None, None),
        ("步骤 2: 用 Wind WSS 函数为每只 IPO 拉横截面字段 (发行价/募资额/认购倍数等)。",
         None, None),
        ("步骤 3: 在 sheet [02_ipo_cornerstones] 手填或半自动填写每只IPO的基石明细 (招股书原文+ticket size)。",
         None, None),
        ("步骤 4: 在 sheet [03_company_fundamentals] 填写公司基本面 (招股书披露的3年财务)。",
         None, None),
        ("步骤 5: 在 sheet [04_ccass_unlocks] 填CCASS解禁日前后的持仓变动。",
         None, None),
        ("步骤 6: sheet [05_hsi_macro] 用 Wind WSD 拉恒指日频; sheet [06_price_index] 列出每只IPO要拉的价格区间, 价格数据保存到独立CSV文件。",
         None, None),
        ("步骤 7: 全部填好后, 把整份 xlsx 发给我, 我来跑 loader 灌库 + 回测。",
         None, None),
        ("", None, None),
        ("颜色规则", Font(name="Arial", size=12, bold=True, color="1F4E78"), None),
        ("• 浅橙底 = 必填字段 (缺失会影响回测有效性)",
         None, FILL_REQUIRED),
        ("• 浅绿底 = 选填字段 (缺失模型用类型先验代偿)",
         None, FILL_OPTIONAL),
        ("• 浅黄底 = 已预填 Wind 公式, 建议放上股票代码后下拖到全部行",
         None, FILL_WIND),
        ("• 浅灰底 = 示例行 (华勤技术 03296.HK), 不要修改", None, FILL_EXAMPLE),
        ("• 蓝色字 = 应该手填的硬编码值; 黑色字 = 公式; 绿色字 = 派生计算",
         None, None),
        ("", None, None),
        ("关键提醒", Font(name="Arial", size=12, bold=True, color="C00000"), None),
        ("⚠ 必须勾选 Wind \"包含已退市/已私有化\" 选项, 否则会引入幸存者偏差。",
         None, None),
        ("⚠ 基石原文名称 (B2 字段) 不要做归一, 招股书原文复制即可, 别名映射由我们处理。",
         None, None),
        ("⚠ 认购金额单位务必统一. 推荐 HKD; 如用USD请在 [02_ipo_cornerstones] sheet 单独标注列。",
         None, None),
        ("⚠ A+H 公司一定用 H 股代码 (如 03296.HK) 不要用 A 股代码 (603296.SH)。",
         None, None),
        ("⚠ 基石名单看招股书最终版, 不是聆讯材料 (基石可能在路演期间增减)。",
         None, None),
        ("", None, None),
        ("Sheet 一览", Font(name="Arial", size=12, bold=True, color="1F4E78"), None),
        ("• 00_README                 — 本说明", None, None),
        ("• 01_ipo_master            — IPO主表 (每只IPO一行, ~250行)",
         None, None),
        ("• 02_ipo_cornerstones      — 基石明细 (长格式, ~2500行)",
         None, None),
        ("• 03_company_fundamentals  — 公司基本面 (每只IPO一行)",
         None, None),
        ("• 04_ccass_unlocks         — CCASS解禁日持仓数据",
         None, None),
        ("• 05_hsi_macro             — 恒指 + 恒科指日频 (1300行)",
         None, None),
        ("• 06_price_index           — 价格采集索引 (告诉你每只IPO拉哪个区间)",
         None, None),
        ("• 90_data_dictionary       — 字段词典", None, None),
        ("• 91_wind_formulas         — Wind 公式速查表", None, None),
    ]

    for i, (text, font, fill) in enumerate(rows, 1):
        c = ws.cell(row=i, column=1, value=text)
        if font:
            c.font = font
        if fill:
            c.fill = fill
        c.alignment = Alignment(horizontal="left", vertical="center",
                                wrap_text=True)
    ws.row_dimensions[1].height = 28


# =============================================================================
# Sheet 2: ipo_master
# =============================================================================

IPO_MASTER_FIELDS = [
    # (字段名, Wind指标, 单位, 必填?, 列宽, 示例值, 是否Wind公式)
    ("stock_code",          "Wind代码",                    "",           True,  16, "03296.HK", False),
    ("company_name_zh",     "证券中文简称",                  "",           True,  20, "华勤技术", False),
    ("listing_date",        "ipo_listdate",                 "yyyy-mm-dd", True,  14, "2026-04-23", "wss"),
    ("pricing_date",        "ipo_setdate",                  "yyyy-mm-dd", True,  14, "2026-04-21", "wss"),
    ("listing_chapter",     "(枚举)",                        "",           True,  18, "a_plus_h", False),
    ("is_a_h",              "AH双重上市",                    "0/1",        True,  10, 1, False),
    ("a_share_code",        "对应A股代码",                   "",           False, 14, "603296.SH", False),
    ("gics_l2",             "行业GICSL2",                    "",           True,  20, "TECH_HARDWARE", "wss"),
    ("offer_price_hkd",     "ipo_price",                     "HKD",        True,  12, 77.70, "wss"),
    ("offer_price_low",     "ipo_offerprice_low",            "HKD",        False, 12, 70.00, "wss"),
    ("offer_price_high",    "ipo_offerprice_high",           "HKD",        False, 12, 77.70, "wss"),
    ("offering_size_hkd",   "ipo_amount",                    "HKD",        True,  16, 4549000000, "wss"),
    ("intl_oversub",        "ipo_intoversub",                "倍",         True,  12, 8.0, "wss"),
    ("public_oversub",      "ipo_puboversub",                "倍",         True,  12, 531.33, "wss"),
    ("clawback_triggered",  "(派生)",                        "0/1",        True,  10, 0, False),
    ("greenshoe_pct",       "ipo_oversinl_pct",              "比例",       False, 10, 0.15, "wss"),
    ("sponsor_primary",     "ipo_underwriter",               "",           True,  28, "中金公司+美银证券", "wss"),
    ("joint_sponsor_count", "(派生:计数+号)",                  "整数",       False, 10, 2, False),
    ("sponsor_tier",        "(枚举1/2/3)",                    "1/2/3",      True,  10, 1, False),
    ("auditor_tier",        "(枚举1/2/3)",                    "1/2/3",      False, 10, 1, False),
    ("pe_at_offer",         "ipo_pe",                        "倍",         False, 10, 12.2, "wss"),
    ("pe_peer_median",      "(人工:5-10家可比)",               "倍",         False, 12, 20.0, False),
    ("last_round_premium",  "(招股书:Pre-IPO最后一轮)",          "比例",       False, 14, -0.28, False),
    ("lockup_months",       "招股书锁定期",                    "月",         False, 10, 6, False),
    ("is_delisted",         "退市状态",                       "0/1",        True,  10, 0, "wss"),
    ("delisting_date",      "退市日期",                       "yyyy-mm-dd", False, 14, "", "wss"),
    ("data_quality_notes",  "(备注)",                         "",           False, 24, "ticket_size estimated", False),
]


def make_ipo_master(wb: Workbook):
    ws = wb.create_sheet("01_ipo_master")
    headers = [f[0] for f in IPO_MASTER_FIELDS]
    widths = [f[4] for f in IPO_MASTER_FIELDS]

    apply_header_row(ws, 1, headers, widths)

    # 子说明行: Wind指标 + 单位 + 必填
    sub = []
    for f in IPO_MASTER_FIELDS:
        wind_ind = f[1]
        unit = f[2]
        req = "★必填" if f[3] else "○选填"
        sub.append(f"{wind_ind}\n{unit}  {req}")
    apply_subheader_row(ws, 2, len(headers), sub)

    # 必填/选填颜色提示
    for i, f in enumerate(IPO_MASTER_FIELDS, 1):
        cell = ws.cell(row=2, column=i)
        cell.fill = FILL_REQUIRED if f[3] else FILL_OPTIONAL

    # 第3行: 华勤示例
    for i, f in enumerate(IPO_MASTER_FIELDS, 1):
        c = ws.cell(row=3, column=i, value=f[5])
        c.fill = FILL_EXAMPLE
        c.border = BORDER_ALL
        c.alignment = ALIGN_LEFT
        c.font = FONT_USER_INPUT
        if f[2] == "HKD":
            c.number_format = '#,##0.00'
        elif f[2] == "比例":
            c.number_format = '0.00%'
        elif f[2] == "yyyy-mm-dd":
            c.number_format = 'yyyy-mm-dd'

    # 第4行: Wind公式模板 (作为文本字符串存储, 避免LibreOffice报错)
    # 用户在 Wind Excel 终端里, 双击单元格删去前导'后会变成真公式
    for i, f in enumerate(IPO_MASTER_FIELDS, 1):
        formula_type = f[6] if len(f) > 6 else False
        c = ws.cell(row=4, column=i)
        if formula_type == "wss":
            # 文本形式存储, 不会被Excel当公式解析(因为前面是引号空格)
            c.value = f'=WSS($A4,"{f[1]}")'
            c.fill = FILL_WIND
            c.font = FONT_FORMULA
            c.number_format = '@'  # 强制文本格式
            c.data_type = 's'      # 字符串类型
        elif i == 1:
            c.value = "<在此填入Wind代码,如 9999.HK>"
            c.font = FONT_NOTE
            c.fill = FILL_WIND
        c.border = BORDER_ALL
        c.alignment = ALIGN_LEFT

    # 数据验证: listing_chapter 下拉
    dv_chapter = DataValidation(
        type="list",
        formula1='"main_board_profitable,a_plus_h,main_board_unprofitable,18a,18c_commercial,18c_precommercial,secondary,spac"',
        allow_blank=True)
    dv_chapter.error = "必须从下拉中选"
    ws.add_data_validation(dv_chapter)
    dv_chapter.add(f"E3:E1000")

    # 数据验证: tier 1/2/3
    dv_tier = DataValidation(type="list", formula1='"1,2,3"', allow_blank=True)
    ws.add_data_validation(dv_tier)
    dv_tier.add("S3:T1000")

    # 数据验证: 0/1
    dv_bool = DataValidation(type="list", formula1='"0,1"', allow_blank=True)
    ws.add_data_validation(dv_bool)
    dv_bool.add("F3:F1000")
    dv_bool.add("O3:O1000")
    dv_bool.add("Y3:Y1000")

    # 冻结首两行 + 第一列
    ws.freeze_panes = "B3"

    # 增加更多空白行供用户填
    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 22


# =============================================================================
# Sheet 3: ipo_cornerstones
# =============================================================================

CORNERSTONE_FIELDS = [
    ("stock_code",            "Wind代码",                "",          True,  14, "03296.HK"),
    ("ipo_listing_date",      "(关联)",                  "yyyy-mm-dd", True,  14, "2026-04-23"),
    ("cornerstone_raw_name",  "招股书原文(不要归一)",       "",          True,  44, "Green Better"),
    ("cornerstone_full_name", "全称/带说明",              "",          False, 44, "Green Better (小米集团 01810.HK 全资子公司)"),
    ("ticket_amount_value",   "认购金额数值",             "",          True,  14, 2000000000),
    ("ticket_amount_ccy",     "币种",                   "HKD/USD",     True,  10, "HKD"),
    ("lockup_months",         "锁定期",                 "月",          True,  8,  6),
    ("affiliation_disclosed", "招股书是否披露关连关系",      "0/1/unknown", False, 10, 1),
    ("affiliation_reason",    "关连原因",                "",          False, 36, "前五大客户"),
    ("data_source",           "数据来源",                "",          False, 24, "prospectus_2026-04-15"),
    ("notes",                 "备注",                   "",          False, 30, ""),
]


# 18家华勤基石作为示范行
HUAQIN_CS_EXAMPLES = [
    ("JPMAMAPL", "JPMAMAPL (J.P. Morgan AM Asia Pac)",            2500e6, "HKD", 6, 0, ""),
    ("UBS AM Singapore", "UBS Asset Management Singapore",          2000e6, "HKD", 6, 0, ""),
    ("上海高毅 + CICC FT swap", "上海高毅资产 (通过CICC FT场外掉期)",     2500e6, "HKD", 6, 0, ""),
    ("Perseverance Asset Management", "Perseverance/兰馨亚洲",     1000e6, "HKD", 6, 0, ""),
    ("泰康人寿", "泰康人寿保险",                                       1500e6, "HKD", 6, 0, ""),
    ("New China Asset Management", "新华资产管理 (新华人寿香港)",        1200e6, "HKD", 6, 0, ""),
    ("光大理财", "光大理财",                                         1000e6, "HKD", 6, 0, ""),
    ("3W Fund", "3W基金",                                          500e6,  "HKD", 6, 0, ""),
    ("Cloud Map", "Cloud Map Capital",                            400e6,  "HKD", 6, 0, ""),
    ("常春藤", "常春藤资本",                                          400e6,  "HKD", 6, 0, ""),
    ("JinYi Capital", "JinYi Capital (代清华教育基金会)",             400e6,  "HKD", 6, 0, ""),
    ("Green Better", "Green Better (小米01810.HK全资)",            2000e6, "HKD", 6, 1, "前五大客户(2023年最大客户)"),
    ("OmniVision HK", "OmniVision HK (豪威/韦尔603501.SH)",        1500e6, "HKD", 6, 1, "CIS图像传感器供应商"),
    ("建滔投资", "建滔投资 (建滔集团00148.HK子公司)",                   1500e6, "HKD", 6, 1, "覆铜板(CCL)供应商"),
    ("宏兴国际", "宏兴国际 (胜宏科技300476.SZ)",                       1000e6, "HKD", 6, 1, "PCB供应商"),
    ("艾唯技术", "艾唯技术 (艾为电子688798.SH)",                      800e6,  "HKD", 6, 1, "模拟/电源管理芯片供应商"),
    ("香港君正", "香港君正 (北京君正300223.SZ)",                      800e6,  "HKD", 6, 1, "存储/MCU芯片供应商"),
    ("Aurora SF", "Aurora SF (晶合集成688249.SH)",                  830e6,  "HKD", 6, 1, "晶圆代工厂, 半导体生态合作方"),
]


def make_cornerstones(wb: Workbook):
    ws = wb.create_sheet("02_ipo_cornerstones")
    headers = [f[0] for f in CORNERSTONE_FIELDS]
    widths = [f[4] for f in CORNERSTONE_FIELDS]

    apply_header_row(ws, 1, headers, widths)

    sub = []
    for f in CORNERSTONE_FIELDS:
        req = "★必填" if f[3] else "○选填"
        sub.append(f"{f[1]}\n{f[2]}  {req}")
    apply_subheader_row(ws, 2, len(headers), sub)

    # 必填/选填颜色提示
    for i, f in enumerate(CORNERSTONE_FIELDS, 1):
        cell = ws.cell(row=2, column=i)
        cell.fill = FILL_REQUIRED if f[3] else FILL_OPTIONAL

    # 华勤18条基石示例
    for j, (raw, full, amt, ccy, lockup, affil, reason) in enumerate(HUAQIN_CS_EXAMPLES):
        row = 3 + j
        ws.cell(row=row, column=1, value="03296.HK")
        ws.cell(row=row, column=2, value="2026-04-23")
        ws.cell(row=row, column=3, value=raw)
        ws.cell(row=row, column=4, value=full)
        ws.cell(row=row, column=5, value=amt).number_format = '#,##0'
        ws.cell(row=row, column=6, value=ccy)
        ws.cell(row=row, column=7, value=lockup)
        ws.cell(row=row, column=8, value=affil)
        ws.cell(row=row, column=9, value=reason)
        ws.cell(row=row, column=10, value="prospectus_2026-04-15")
        for c in range(1, len(CORNERSTONE_FIELDS) + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = FILL_EXAMPLE
            cell.border = BORDER_ALL
            cell.font = FONT_USER_INPUT

    # 数据验证
    dv_ccy = DataValidation(type="list", formula1='"HKD,USD,CNY"',
                            allow_blank=True)
    ws.add_data_validation(dv_ccy)
    dv_ccy.add("F3:F5000")
    dv_affil = DataValidation(type="list", formula1='"0,1,unknown"',
                              allow_blank=True)
    ws.add_data_validation(dv_affil)
    dv_affil.add("H3:H5000")

    ws.freeze_panes = "B3"


# =============================================================================
# Sheet 4: company_fundamentals
# =============================================================================

FUNDAMENTAL_FIELDS = [
    ("stock_code",            "Wind代码",                  "",   True,  14,  "03296.HK"),
    ("revenue_n3",            "招股书披露 N-3 年营收",       "RMB亿", True, 14, 853.4),
    ("revenue_n2",            "招股书披露 N-2 年营收",       "RMB亿", True, 14, 1098.8),
    ("revenue_n1",            "招股书披露 N-1 年营收",       "RMB亿", True, 14, 1714.4),
    ("net_profit_n3",         "N-3 年净利润",              "RMB亿", True, 14, 26.57),
    ("net_profit_n2",         "N-2 年净利润",              "RMB亿", True, 14, 29.16),
    ("net_profit_n1",         "N-1 年净利润",              "RMB亿", True, 14, 41.31),
    ("gross_margin_n3",       "N-3 年毛利率",              "%",     True, 12, 0.109),
    ("gross_margin_n2",       "N-2 年毛利率",              "%",     True, 12, 0.093),
    ("gross_margin_n1",       "N-1 年毛利率",              "%",     True, 12, 0.077),
    ("roe_avg_3y",            "3年平均ROE",                "%",     False, 12, 0.20),
    ("net_debt_to_ebitda",    "净负债/EBITDA (N-1)",        "倍",    False, 12, 1.0),
    ("fcf_positive_years_3y", "近3年自由现金流为正的年数",     "0-3",   False, 14, 1),
    ("top1_customer_pct_n1",  "N-1 年第一大客户收入占比",     "%",     False, 14, 0.149),
    ("top5_customer_pct_n1",  "N-1 年前五大客户收入占比",     "%",     False, 14, 0.541),
    ("revenue_cagr_3y",       "(派生:loader自动算)",         "%",     False, 14, 0.418),
    ("gross_margin_trend",    "(派生:loader自动算)",         "",      False, 14, -0.016),
]


def make_fundamentals(wb: Workbook):
    ws = wb.create_sheet("03_company_fundamentals")
    headers = [f[0] for f in FUNDAMENTAL_FIELDS]
    widths = [f[4] for f in FUNDAMENTAL_FIELDS]

    apply_header_row(ws, 1, headers, widths)

    sub = []
    for f in FUNDAMENTAL_FIELDS:
        req = "★必填" if f[3] else "○选填"
        sub.append(f"{f[1]}\n{f[2]}  {req}")
    apply_subheader_row(ws, 2, len(headers), sub)
    for i, f in enumerate(FUNDAMENTAL_FIELDS, 1):
        cell = ws.cell(row=2, column=i)
        cell.fill = FILL_REQUIRED if f[3] else FILL_OPTIONAL

    # 华勤示例
    for i, f in enumerate(FUNDAMENTAL_FIELDS, 1):
        c = ws.cell(row=3, column=i, value=f[5])
        c.fill = FILL_EXAMPLE
        c.border = BORDER_ALL
        c.alignment = ALIGN_LEFT
        c.font = FONT_USER_INPUT
        if "%" in f[2]:
            c.number_format = '0.0%'
        elif f[2] == "RMB亿":
            c.number_format = '#,##0.00'

    ws.freeze_panes = "B3"


# =============================================================================
# Sheet 5: ccass_unlocks
# =============================================================================

CCASS_FIELDS = [
    ("stock_code",                  "Wind代码",                  "",            True,  14, "03296.HK"),
    ("ipo_listing_date",            "(关联)",                    "yyyy-mm-dd",   True,  14, "2026-04-23"),
    ("unlock_date",                 "锁定期解禁日 (上市+6个月)",     "yyyy-mm-dd",   True,  14, "2026-10-23"),
    ("ccass_pct_basestone_d_minus30","解禁前30天 基石持仓占比",     "%",            False, 18, ""),
    ("ccass_pct_basestone_d0",      "解禁日 基石持仓占比",         "%",            False, 16, ""),
    ("ccass_pct_basestone_d_plus30","解禁后30天 基石持仓占比",     "%",            False, 18, ""),
    ("ccass_pct_basestone_d_plus90","解禁后90天 基石持仓占比",     "%",            False, 18, ""),
    ("price_d0_unlock",             "解禁日收盘价",              "HKD",           False, 12, ""),
    ("price_d_plus30",              "解禁+30天收盘价",           "HKD",           False, 12, ""),
    ("price_d_plus90",              "解禁+90天收盘价",           "HKD",           False, 12, ""),
    ("notes",                       "备注",                     "",             False, 24, ""),
]


def make_ccass(wb: Workbook):
    ws = wb.create_sheet("04_ccass_unlocks")
    headers = [f[0] for f in CCASS_FIELDS]
    widths = [f[4] for f in CCASS_FIELDS]

    apply_header_row(ws, 1, headers, widths)

    sub = []
    for f in CCASS_FIELDS:
        req = "★必填" if f[3] else "○选填"
        sub.append(f"{f[1]}\n{f[2]}  {req}")
    apply_subheader_row(ws, 2, len(headers), sub)
    for i, f in enumerate(CCASS_FIELDS, 1):
        cell = ws.cell(row=2, column=i)
        cell.fill = FILL_REQUIRED if f[3] else FILL_OPTIONAL

    for i, f in enumerate(CCASS_FIELDS, 1):
        c = ws.cell(row=3, column=i, value=f[5])
        c.fill = FILL_EXAMPLE
        c.border = BORDER_ALL
        c.alignment = ALIGN_LEFT
        c.font = FONT_USER_INPUT

    # 提示行
    ws.cell(row=4, column=1, value="说明: CCASS数据可用 Wind 函数 ccass_pct(stockcode, date), 或导出 HKEX 官网 CCASS Stock Tracker 数据.")
    ws.cell(row=4, column=1).font = FONT_NOTE
    ws.merge_cells("A4:K4")

    ws.freeze_panes = "B3"


# =============================================================================
# Sheet 6: hsi_macro
# =============================================================================

def make_hsi_macro(wb: Workbook):
    ws = wb.create_sheet("05_hsi_macro")
    headers = ["trade_date", "hsi_close", "hscei_close", "hsi_volume_hkd",
               "hibor_3m", "southbound_net_flow_hkd"]
    widths = [14, 14, 14, 18, 12, 22]
    apply_header_row(ws, 1, headers, widths)

    sub_info = [
        "yyyy-mm-dd\n★必填", "Wind: HSI.HI close\n★必填",
        "Wind: HSCEI.HI close\n★必填",
        "成交额\n○选填", "Wind: HIBOR3M.IR\n○选填",
        "Wind: 南向资金净流入\n○选填",
    ]
    apply_subheader_row(ws, 2, len(headers), sub_info)
    for i in range(1, len(headers) + 1):
        cell = ws.cell(row=2, column=i)
        cell.fill = FILL_REQUIRED if i <= 3 else FILL_OPTIONAL

    # 第3行: Wind公式示例 (作为文本字符串存储)
    ws.cell(row=3, column=1, value="2022-01-04")
    ws.cell(row=3, column=1).font = FONT_USER_INPUT
    formulas = [
        '=WSD("HSI.HI","close",$A3,$A3,"")',
        '=WSD("HSCEI.HI","close",$A3,$A3,"")',
        '=WSD("HSI.HI","amt",$A3,$A3,"")',
        '=WSD("HIBOR3M.IR","close",$A3,$A3,"")',
        '=WSET("nethkscconnect","date="&$A3)',
    ]
    for col_i, formula in enumerate(formulas, 2):
        c = ws.cell(row=3, column=col_i, value=formula)
        c.fill = FILL_WIND
        c.number_format = '@'
        c.data_type = 's'
        c.font = FONT_FORMULA

    for i in range(1, len(headers) + 1):
        ws.cell(row=3, column=i).border = BORDER_ALL

    # 提示
    ws.cell(row=4, column=1,
            value="提示: 用 Wind WSD 一次性拉所有日期范围更高效, 例如:")
    c5 = ws.cell(row=5, column=1,
            value='=WSD("HSI.HI","close,amt","2022-01-01","2026-04-30","")')
    c5.number_format = '@'
    c5.data_type = 's'
    ws.cell(row=4, column=1).font = FONT_NOTE
    c5.font = FONT_NOTE
    ws.merge_cells("A4:F4")
    ws.merge_cells("A5:F5")

    ws.freeze_panes = "B3"


# =============================================================================
# Sheet 7: price_index (告诉用户每只IPO拉哪段价格)
# =============================================================================

def make_price_index(wb: Workbook):
    ws = wb.create_sheet("06_price_index")
    headers = ["stock_code", "listing_date", "price_window_start",
               "price_window_end", "csv_filename", "wind_formula_template"]
    widths = [14, 14, 18, 18, 28, 56]
    apply_header_row(ws, 1, headers, widths)

    sub_info = [
        "Wind代码\n★必填", "上市日(从sheet01拷过来)\n★必填",
        "(派生:上市前5天)\n", "(派生:上市后400天)\n", "(派生)\n",
        "(派生)\n",
    ]
    apply_subheader_row(ws, 2, len(headers), sub_info)

    # 华勤示例 (用 hardcode 值, 不用公式以避免 #NAME 错误)
    ws.cell(row=3, column=1, value="03296.HK").font = FONT_USER_INPUT
    ws.cell(row=3, column=2, value="2026-04-23").font = FONT_USER_INPUT
    ws.cell(row=3, column=2).number_format = "yyyy-mm-dd"
    ws.cell(row=3, column=3, value="2026-04-18").font = FONT_DERIVED
    ws.cell(row=3, column=3).number_format = "yyyy-mm-dd"
    ws.cell(row=3, column=4, value="2027-05-28").font = FONT_DERIVED
    ws.cell(row=3, column=4).number_format = "yyyy-mm-dd"
    ws.cell(row=3, column=5, value="03296_HK.csv").font = FONT_DERIVED
    # Wind 公式作为文本字符串
    c6 = ws.cell(row=3, column=6,
                 value='=WSD("03296.HK","close,amt,trade_status","2026-04-18","2027-05-28","")')
    c6.font = FONT_FORMULA
    c6.number_format = '@'
    c6.data_type = 's'

    for i in range(1, 7):
        ws.cell(row=3, column=i).fill = FILL_EXAMPLE
        ws.cell(row=3, column=i).border = BORDER_ALL

    # 说明
    ws.cell(row=5, column=1,
            value="工作流: 把 sheet01 的全部 (stock_code, listing_date) 拷贝到 A/B 列, "
                  "F列会自动生成 Wind 公式. 把每个公式输出粘贴到独立 CSV 文件, "
                  "文件名用 E 列的命名规则.")
    ws.cell(row=5, column=1).font = FONT_NOTE
    ws.merge_cells("A5:F5")
    ws.row_dimensions[5].height = 36

    ws.freeze_panes = "B3"


# =============================================================================
# Sheet 8: data_dictionary
# =============================================================================

DICT_ROWS = [
    ("listing_chapter", "枚举", "main_board_profitable / a_plus_h / main_board_unprofitable / 18a / 18c_commercial / 18c_precommercial / secondary / spac"),
    ("sponsor_tier", "枚举", "1=中金/MS/GS/UBS/JPM/华泰国际; 2=海通国际/招银国际/农银国际/建银国际; 3=其他"),
    ("auditor_tier", "枚举", "1=PwC/EY/KPMG/Deloitte 四大; 2=本地大所; 3=其他"),
    ("affiliation_disclosed", "枚举", "1=招股书披露与发行人/控股股东/保荐人有关连; 0=明确无关连; unknown=招股书未提"),
    ("ticket_amount_ccy", "枚举", "HKD / USD / CNY (USD和CNY会按招股日中间价转HKD)"),
    ("intl_oversub", "数值", "国际配售认购倍数. <1.5x 会触发 Layer 1 否决条款"),
    ("public_oversub", "数值", "公开发售认购倍数. 单看不重要, 但与intl严重背离时(如>100x但intl<3x)是危险信号"),
    ("clawback_triggered", "布尔", "机制A可能触发(超额时), 机制B永远False"),
    ("greenshoe_pct", "数值", "绿鞋占发行规模比例, 通常0.15"),
    ("last_round_premium", "数值", "(发行估值/Pre-IPO最后一轮估值)-1. A+H公司常用 H股发行价/A股价格-1"),
    ("lockup_months", "数值", "基石锁定期(月), 通常6, 部分项目分批解锁取最长"),
    ("is_delisted", "布尔", "回测时不剔除, 用于反幸存者偏差"),
]


def make_data_dictionary(wb: Workbook):
    ws = wb.create_sheet("90_data_dictionary")
    apply_header_row(ws, 1, ["字段名", "类型", "取值说明"],
                     [28, 12, 90])
    for i, r in enumerate(DICT_ROWS, 2):
        for j, v in enumerate(r, 1):
            c = ws.cell(row=i, column=j, value=v)
            c.font = Font(name="Arial", size=10)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            c.border = BORDER_ALL
        ws.row_dimensions[i].height = 32


# =============================================================================
# Sheet 9: wind_formulas
# =============================================================================

WIND_FORMULA_ROWS = [
    # (用途, Wind公式, 字段示例, 说明)
    ("拉 2022-2026/4 全部HK主板IPO列表",
     '=WSET("newstock","startdate=2022-01-01;enddate=2026-04-30;exchange=hkex;sectorcode=a201020700000000")',
     "ipo_master.stock_code",
     "**先在终端运行此公式, 把返回的代码列表粘贴到 sheet01 的 A 列**"),
    ("是否包含已退市", "勾选Wind函数选项 includesuspended=1",
     "is_delisted", "右键函数→选项→包含已退市"),
    ("单只IPO横截面字段", '=WSS("03296.HK","ipo_listdate,ipo_setdate,ipo_price,ipo_amount,ipo_intoversub,ipo_puboversub,ipo_pe")',
     "多个字段一次拉", "推荐做法: 把整列字段名用逗号串联, 一次性拉"),
    ("发行价", '=WSS($A3,"ipo_price")', "offer_price_hkd", ""),
    ("募资总额", '=WSS($A3,"ipo_amount")', "offering_size_hkd", "通常HKD口径, 但要确认单位"),
    ("国际配售倍数", '=WSS($A3,"ipo_intoversub")', "intl_oversub",
     "如显示空白, 用配售结果公告补"),
    ("公开发售倍数", '=WSS($A3,"ipo_puboversub")', "public_oversub", ""),
    ("发行PE", '=WSS($A3,"ipo_pe")', "pe_at_offer", ""),
    ("保荐人", '=WSS($A3,"ipo_underwriter")', "sponsor_primary", ""),
    ("行业GICS", '=WSS($A3,"sec_gicsindustryname2")', "gics_l2", ""),
    ("退市状态", '=WSS($A3,"sec_status")', "is_delisted",
     '返回值映射: "退市"->1, "正常"->0'),
    ("退市日期", '=WSS($A3,"delist_date")', "delisting_date", ""),
    ("基石明细", "Wind 没有现成函数",
     "ipo_cornerstones",
     "需手工从招股书中复制. 招股书在 HKEX 披露易 (https://www.hkexnews.hk) 下载. 推荐用Adobe Acrobat或Excel粘贴板提取"),
    ("CCASS 持仓占比", '=WSS("03296.HK","ccass_pct;tradeDate=2026-10-23")',
     "ccass_pct_*", "需要参数 tradeDate"),
    ("恒指日频", '=WSD("HSI.HI","close,amt","2022-01-01","2026-04-30","")',
     "05_hsi_macro", "一次拉全期间"),
    ("HIBOR 3M", '=WSD("HIBOR3M.IR","close","2022-01-01","2026-04-30","")',
     "hibor_3m", "选填, 用于资金成本"),
    ("南向资金", '=WSD("AHKSCH.WI","close","2022-01-01","2026-04-30","")',
     "southbound_net_flow", "南向资金累计净流入指数"),
    ("单只IPO价格历史", '=WSD("03296.HK","close,amt,trade_status","2026-04-18","2027-05-28","")',
     "06_price_index", "上市前5天到上市后400天"),
]


def make_wind_formulas(wb: Workbook):
    ws = wb.create_sheet("91_wind_formulas")
    headers = ["用途", "Wind公式", "目标字段", "说明"]
    widths = [32, 80, 28, 70]
    apply_header_row(ws, 1, headers, widths)

    for i, r in enumerate(WIND_FORMULA_ROWS, 2):
        for j, v in enumerate(r, 1):
            c = ws.cell(row=i, column=j, value=v)
            c.font = Font(name="Arial", size=10,
                          color="C00000" if j == 2 else "000000")
            c.alignment = Alignment(wrap_text=True, vertical="top",
                                    horizontal="left")
            c.border = BORDER_ALL
            if j == 2:
                # Wind公式列设为文本格式, 避免 #NAME 错误
                c.number_format = '@'
                c.data_type = 's'
        ws.row_dimensions[i].height = 36

    # 顶部提示
    ws.merge_cells("A2:D2")
    ws.row_dimensions[2].height = 0  # 隐藏


# =============================================================================
# Main
# =============================================================================

def build_template(out_path: str):
    wb = Workbook()
    # 删除默认的Sheet
    wb.remove(wb.active)

    make_readme(wb)
    make_ipo_master(wb)
    make_cornerstones(wb)
    make_fundamentals(wb)
    make_ccass(wb)
    make_hsi_macro(wb)
    make_price_index(wb)
    make_data_dictionary(wb)
    make_wind_formulas(wb)

    wb.save(out_path)
    print(f"模板已生成: {out_path}")


if __name__ == "__main__":
    import sys
    out = sys.argv[1] if len(sys.argv) > 1 else "/tmp/nacs_data_template.xlsx"
    build_template(out)
